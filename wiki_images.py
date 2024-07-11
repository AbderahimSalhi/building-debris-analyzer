"""
Building Debris Image Analyzer
Author: Abderahim Salhi
Version: 3.0
date: 1/5/2024

This script scrapes images of building debris from Wikimedia Commons,
analyzes them using machine learning models (CLIP and YOLOv5),
and generates a labeled dataset in Excel format.
"""

import requests
import os
import json
import time
from openpyxl import Workbook
from openpyxl.styles import Font
import torch
from transformers import CLIPProcessor, CLIPModel
from PIL import Image, UnidentifiedImageError

# Check CUDA availability
print(f"CUDA available: {torch.cuda.is_available()}")
print(f"CUDA version: {torch.version.cuda}")
print(f"Number of CUDA devices: {torch.cuda.device_count()}")

# Try to import YOLOv5, fall back to CLIP-only if not available
try:
    import yolov5
except ImportError:
    print("YOLOv5 not found. Falling back to CLIP-only analysis.")
    yolov5 = None

# User agent for API requests
USER_AGENT = "BuildingDebrisImageBot/1.0 (https://github.com/yourusername/your-repo; youremail@example.com) python-requests/2.26.0"

class ImageLabeler:
    def __init__(self):
        # Initialize disaster types and data storage
        self.disaster_types = ["tornado", "hurricane", "wildfire", "earthquake", "flood", "tsunami"]
        self.images_data = []

        # Initialize CLIP model and processor
        self.clip_processor = CLIPProcessor.from_pretrained("openai/clip-vit-base-patch32")
        self.clip_model = CLIPModel.from_pretrained("openai/clip-vit-base-patch32")

        # Set up device (GPU if available, else CPU)
        self.device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
        print(f"Using device: {self.device}")

        # Try to load YOLOv5 model
        try:
            self.yolo_model = torch.hub.load('ultralytics/yolov5', 'yolov5s', pretrained=True)
            self.yolo_model.to(self.device)
            print("YOLOv5 model loaded successfully")
        except Exception as e:
            print(f"Error loading YOLOv5 model: {e}")
            print("Falling back to CLIP-only analysis")
            self.yolo_model = None

        # Move CLIP model to GPU if available
        self.clip_model.to(self.device)

    def analyze_image(self, image_path):
        # Open and preprocess the image
        try:
            image = Image.open(image_path)
            image.verify()  # Verify that it's a valid image
            image = Image.open(image_path)  # Reopen the image after verification
            if image.mode != 'RGB':
                image = image.convert('RGB')
            image_array = np.array(image)
        except (OSError, UnidentifiedImageError, ValueError) as e:
            print(f"Error processing image {image_path}: {e}")
            return "unknown", "unknown"

        # Analyze the image using CLIP and YOLOv5 (if available)
        try:
            # Use CLIP to determine disaster type
            inputs = self.clip_processor(text=self.disaster_types, images=image_array, return_tensors="pt", padding=True)
            inputs = {k: v.to(self.device) for k, v in inputs.items()}
            with torch.no_grad():
                outputs = self.clip_model(**inputs)
            logits_per_image = outputs.logits_per_image
            probs = logits_per_image.softmax(dim=1)
            disaster_type = self.disaster_types[probs.argmax().item()]

            # Detect human presence using YOLOv5 or CLIP
            if self.yolo_model:
                results = self.yolo_model(image_array)
                humans_detected = any(cls == 0 for *box, conf, cls in results.xyxy[0].cpu().numpy())
                humans_visible = "yes" if humans_detected else "no"
            else:
                human_inputs = self.clip_processor(text=["a photo with people", "a photo without people"], images=image_array, return_tensors="pt", padding=True)
                human_inputs = {k: v.to(self.device) for k, v in human_inputs.items()}
                with torch.no_grad():
                    human_outputs = self.clip_model(**human_inputs)
                human_probs = human_outputs.logits_per_image.softmax(dim=1)
                humans_visible = "yes" if human_probs[0][0] > human_probs[0][1] else "no"

            return disaster_type, humans_visible
        except Exception as e:
            print(f"Error analyzing image {image_path}: {e}")
            return "unknown", "unknown"

    def label_image(self, filename, description):
        # Process and label a single image
        image_path = os.path.join("building_debris_images", filename)
        if not self.is_image_file(image_path):
            print(f"Skipping non-image file: {filename}")
            return

        disaster_type, humans_visible = self.analyze_image(image_path)
        building_intact = "unknown"
        if "collapsed" in description.lower() or "damaged" in description.lower():
            building_intact = "no"
        elif "intact" in description.lower():
            building_intact = "yes"

        self.images_data.append({
            "filename": filename,
            "disaster_type": disaster_type,
            "building_intact": building_intact,
            "humans_visible": humans_visible,
            "description": description
        })

    def is_image_file(self, filepath):
        # Check if the file is a valid image based on its extension
        image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.webp']
        _, extension = os.path.splitext(filepath)
        return extension.lower() in image_extensions

    def generate_excel(self, output_file):
        # Generate an Excel file with the labeled data
        wb = Workbook()
        ws = wb.active
        ws.title = "Image Labels"

        headers = ["Filename", "Disaster Type", "Building Intact", "Humans Visible", "Description"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        for row, data in enumerate(self.images_data, start=2):
            ws.cell(row=row, column=1, value=data["filename"])
            ws.cell(row=row, column=2, value=data["disaster_type"])
            ws.cell(row=row, column=3, value=data["building_intact"])
            ws.cell(row=row, column=4, value=data["humans_visible"])
            ws.cell(row=row, column=5, value=data["description"])

        wb.save(output_file)
        print(f"Excel file generated: {output_file}")



# """
# Building Debris Image Analyzer
# Author: Abderahim Salhi
# Version: 1.0
# date: 12/28/2023

# This script scrapes images of building debris from Wikimedia Commons,
# analyzes them using machine learning models (CLIP and YOLOv5),
# and generates a labeled dataset in Excel format.
# """
# import requests
# import os
# import json
# import time

# # Define a user agent string
# USER_AGENT = "BuildingDebrisImageBot/1.0 (https://github.com/yourusername/your-repo; youremail@example.com) python-requests/2.26.0"

# def download_image(url, folder, retries=3):
#     headers = {'User-Agent': USER_AGENT}
#     for attempt in range(retries):
#         try:
#             response = requests.get(url, headers=headers, timeout=10)
#             response.raise_for_status()
#             filename = url.split('/')[-1]
#             filepath = os.path.join(folder, filename)
#             with open(filepath, 'wb') as f:
#                 f.write(response.content)
#             print(f"Downloaded: {filename}")
#             return True
#         except requests.exceptions.RequestException as e:
#             print(f"Attempt {attempt + 1} failed: {e}")
#             time.sleep(1)
#     print(f"Failed to download after {retries} attempts: {url}")
#     return False

# def search_wikimedia_commons(search_term, num_images=20):
#     api_url = "https://commons.wikimedia.org/w/api.php"
#     params = {
#         "action": "query",
#         "format": "json",
#         "list": "search",
#         "srsearch": search_term,
#         "srnamespace": "6",
#         "srlimit": num_images,
#         "srprop": "size|wordcount|timestamp|snippet"
#     }
#     headers = {'User-Agent': USER_AGENT}
    
#     try:
#         response = requests.get(api_url, params=params, headers=headers, timeout=10)
#         response.raise_for_status()
#         data = response.json()
#         return data['query']['search']
#     except requests.exceptions.RequestException as e:
#         print(f"API request failed: {e}")
#         return []

# def get_image_url(title):
#     api_url = "https://commons.wikimedia.org/w/api.php"
#     params = {
#         "action": "query",
#         "format": "json",
#         "titles": title,
#         "prop": "imageinfo",
#         "iiprop": "url"
#     }
#     headers = {'User-Agent': USER_AGENT}
    
#     try:
#         response = requests.get(api_url, params=params, headers=headers, timeout=10)
#         response.raise_for_status()
#         data = response.json()
#         pages = data['query']['pages']
#         for page_id in pages:
#             if 'imageinfo' in pages[page_id]:
#                 return pages[page_id]['imageinfo'][0]['url']
#     except requests.exceptions.RequestException as e:
#         print(f"Failed to get image URL: {e}")
#     return None

# def scrape_wikimedia_commons(search_term, num_images=20):
#     download_folder = "building_debris_images"
#     os.makedirs(download_folder, exist_ok=True)
    
#     search_results = search_wikimedia_commons(search_term, num_images)
    
#     successful_downloads = 0
#     for i, result in enumerate(search_results):
#         title = result['title']
#         image_url = get_image_url(title)
#         if image_url:
#             if download_image(image_url, download_folder):
#                 successful_downloads += 1
#         else:
#             print(f"Couldn't find image URL for: {title}")
        
#         if successful_downloads >= num_images:
#             break

#     print(f"Successfully downloaded {successful_downloads} images.")

# if __name__ == "__main__":
#     search_term = "building debris damage"
#     scrape_wikimedia_commons(search_term, num_images=20)


#Building Debris Image Analyzer
# Author: Abderahim Salhi
# Version: 2.0
# date: 1/4/2024

# This script scrapes images of building debris from Wikimedia Commons,
# analyzes them using machine learning models (CLIP and YOLOv5),
# and generates a labeled dataset in Excel format.
# """
# import requests
# import os
# import json
# import time
# from openpyxl import Workbook
# from openpyxl.styles import Font

# USER_AGENT = "BuildingDebrisImageBot/1.0 (https://github.com/yourusername/your-repo; youremail@example.com) python-requests/2.26.0"

# class ImageLabeler:
#     def __init__(self):
#         self.disaster_types = ["tornado", "hurricane", "wildfire", "earthquake", "flood", "tsunami"]
#         self.images_data = []

#     def label_image(self, filename, description):
#         disaster_type = "unknown"
#         for d_type in self.disaster_types:
#             if d_type in description.lower():
#                 disaster_type = d_type
#                 break
        
#         building_intact = "unknown"
#         if "collapsed" in description.lower() or "damaged" in description.lower():
#             building_intact = "no"
#         elif "intact" in description.lower():
#             building_intact = "yes"
        
#         humans_visible = "unknown"
#         if "people" in description.lower() or "person" in description.lower():
#             humans_visible = "yes"
        
#         self.images_data.append({
#             "filename": filename,
#             "disaster_type": disaster_type,
#             "building_intact": building_intact,
#             "humans_visible": humans_visible,
#             "description": description
#         })

#     def generate_excel(self, output_file):
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Image Labels"

#         headers = ["Filename", "Disaster Type", "Building Intact", "Humans Visible", "Description"]
#         for col, header in enumerate(headers, start=1):
#             cell = ws.cell(row=1, column=col, value=header)
#             cell.font = Font(bold=True)

#         for row, data in enumerate(self.images_data, start=2):
#             ws.cell(row=row, column=1, value=data["filename"])
#             ws.cell(row=row, column=2, value=data["disaster_type"])
#             ws.cell(row=row, column=3, value=data["building_intact"])
#             ws.cell(row=row, column=4, value=data["humans_visible"])
#             ws.cell(row=row, column=5, value=data["description"])

#         wb.save(output_file)
#         print(f"Excel file generated: {output_file}")

# def download_image(url, folder, labeler):
#     headers = {'User-Agent': USER_AGENT}
#     for attempt in range(3):
#         try:
#             response = requests.get(url, headers=headers, timeout=10)
#             response.raise_for_status()
#             filename = url.split('/')[-1]
#             filepath = os.path.join(folder, filename)
#             with open(filepath, 'wb') as f:
#                 f.write(response.content)
#             print(f"Downloaded: {filename}")
#             return filename
#         except requests.exceptions.RequestException as e:
#             print(f"Attempt {attempt + 1} failed: {e}")
#             time.sleep(1)
#     print(f"Failed to download after 3 attempts: {url}")
#     return None

# def search_wikimedia_commons(search_term, num_images=20):
#     api_url = "https://commons.wikimedia.org/w/api.php"
#     params = {
#         "action": "query",
#         "format": "json",
#         "list": "search",
#         "srsearch": search_term,
#         "srnamespace": "6",
#         "srlimit": num_images,
#         "srprop": "size|wordcount|timestamp|snippet"
#     }
#     headers = {'User-Agent': USER_AGENT}
    
#     try:
#         response = requests.get(api_url, params=params, headers=headers, timeout=10)
#         response.raise_for_status()
#         data = response.json()
#         return data['query']['search']
#     except requests.exceptions.RequestException as e:
#         print(f"API request failed: {e}")
#         return []

# def get_image_url(title):
#     api_url = "https://commons.wikimedia.org/w/api.php"
#     params = {
#         "action": "query",
#         "format": "json",
#         "titles": title,
#         "prop": "imageinfo",
#         "iiprop": "url|extmetadata"
#     }
#     headers = {'User-Agent': USER_AGENT}
    
#     try:
#         response = requests.get(api_url, params=params, headers=headers, timeout=10)
#         response.raise_for_status()
#         data = response.json()
#         pages = data['query']['pages']
#         for page_id in pages:
#             if 'imageinfo' in pages[page_id]:
#                 image_info = pages[page_id]['imageinfo'][0]
#                 return image_info['url'], image_info['extmetadata'].get('ImageDescription', {}).get('value', '')
#     except requests.exceptions.RequestException as e:
#         print(f"Failed to get image URL: {e}")
#     return None, None

# def scrape_wikimedia_commons(search_term, num_images=20):
#     download_folder = "building_debris_images"
#     os.makedirs(download_folder, exist_ok=True)
    
#     search_results = search_wikimedia_commons(search_term, num_images)
    
#     labeler = ImageLabeler()
#     successful_downloads = 0
#     for i, result in enumerate(search_results):
#         title = result['title']
#         image_url, description = get_image_url(title)
#         if image_url:
#             filename = download_image(image_url, download_folder, labeler)
#             if filename:
#                 labeler.label_image(filename, description)
#                 successful_downloads += 1
#         else:
#             print(f"Couldn't find image URL for: {title}")
        
#         if successful_downloads >= num_images:
#             break

#     print(f"Successfully downloaded {successful_downloads} images.")
#     labeler.generate_excel("image_labels.xlsx")

# if __name__ == "__main__":
#     search_term = "building debris damage"
#     scrape_wikimedia_commons(search_term, num_images=20)

# class ImageLabeler:
#     def __init__(self):
#         self.disaster_types = ["tornado", "hurricane", "wildfire", "earthquake", "flood", "tsunami"]
#         self.images_data = []
#         self.model = CLIPModel.from_pretrained("openai/clip-vit-base-patch32")
#         self.processor = CLIPProcessor.from_pretrained("openai/clip-vit-base-patch32")

#     def analyze_image(self, image_path):
#         image = Image.open(image_path)
#         inputs = self.processor(text=self.disaster_types, images=image, return_tensors="pt", padding=True)
#         outputs = self.model(**inputs)
#         logits_per_image = outputs.logits_per_image
#         probs = logits_per_image.softmax(dim=1)
#         return self.disaster_types[probs.argmax().item()]

#     def label_image(self, filename, description):
#         image_path = os.path.join("building_debris_images", filename)
#         disaster_type = self.analyze_image(image_path)
        
#         building_intact = "unknown"
#         if "collapsed" in description.lower() or "damaged" in description.lower():
#             building_intact = "no"
#         elif "intact" in description.lower():
#             building_intact = "yes"
        
#         humans_visible = "unknown"
#         if "people" in description.lower() or "person" in description.lower():
#             humans_visible = "yes"
        
#         self.images_data.append({
#             "filename": filename,
#             "disaster_type": disaster_type,
#             "building_intact": building_intact,
#             "humans_visible": humans_visible,
#             "description": description
#         })

#     def generate_excel(self, output_file):
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Image Labels"

#         headers = ["Filename", "Disaster Type", "Building Intact", "Humans Visible", "Description"]
#         for col, header in enumerate(headers, start=1):
#             cell = ws.cell(row=1, column=col, value=header)
#             cell.font = Font(bold=True)

#         for row, data in enumerate(self.images_data, start=2):
#             ws.cell(row=row, column=1, value=data["filename"])
#             ws.cell(row=row, column=2, value=data["disaster_type"])
#             ws.cell(row=row, column=3, value=data["building_intact"])
#             ws.cell(row=row, column=4, value=data["humans_visible"])
#             ws.cell(row=row, column=5, value=data["description"])

#         wb.save(output_file)
#         print(f"Excel file generated: {output_file}")

